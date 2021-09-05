import urllib.request as req
from bs4 import BeautifulSoup
import os
import openpyxl
import datetime
from openpyxl.drawing.image import Image

# 이미지 저장할 폴더 생성
if not os.path.exists("./멜론이미지"):
    os.mkdir("./멜론이미지")

header = req.Request("https://www.melon.com/chart/index.htm", headers={"User-Agent":"Mozilla/5.0"})
code = req.urlopen(header)
soup = BeautifulSoup(code, "html.parser")
title = soup.select("div.ellipsis.rank01 > span > a")
name = soup.select("div.ellipsis.rank02 > span")
album = soup.select("div.ellipsis.rank03 > a")
img = soup.select("a.image_typeAll > img")

# 엑셀 파일 생성
if not os.path.exists("./멜론_크롤링.xlsx"):
    openpyxl.Workbook().save("./멜론_크롤링.xlsx")

# 엑셀 파일 불러오기
book = openpyxl.load_workbook("./멜론_크롤링.xlsx")
# 쓸데 없는 시트 지우기
if "Sheet" in book.sheetnames:
    book.remove(book["Sheet"])
sheet = book.create_sheet()
now = datetime.datetime.now()
sheet.title = f"{now.year}년 {now.month}월 {now.day}일 {now.hour}시 {now.minute}분 {now.second}초"
# 열 너비 조절
sheet.column_dimensions["A"].width = 15
sheet.column_dimensions["B"].width = 50
sheet.column_dimensions["C"].width = 30
sheet.column_dimensions["D"].width = 50

for i in range(len(title)):
    img_file_name = f"./멜론이미지/{i+1}.png"
    req.urlretrieve(img[i].attrs["src"], img_file_name)
    print("{}위. {} - {}".format(i+1, title[i].string, name[i].text))
    img_for_excel = Image(img_file_name)
    sheet.add_image(img_for_excel, f"A{i+1}")
    sheet.cell(row=i + 1, column=2).value = title[i].string
    sheet.cell(row=i + 1, column=3).value = name[i].text
    sheet.cell(row=i + 1, column=4).value = album[i].string
    sheet.row_dimensions[i+1].height = 90
    book.save("./멜론_크롤링.xlsx")
